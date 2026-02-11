#!/usr/bin/env python3
"""
ScrapMonster Automotive Yard Scraper
Scrapes automotive yard listings from ScrapMonster and exports to Excel.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys
from typing import List, Dict, Optional


class ScrapMonsterScraper:
    """Scraper for ScrapMonster automotive yard listings."""
    
    def __init__(self, url: str):
        """
        Initialize the scraper.
        
        Args:
            url: The URL to scrape
        """
        self.url = url
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        self.listings = []
    
    def fetch_page(self) -> Optional[BeautifulSoup]:
        """
        Fetch and parse the webpage.
        
        Returns:
            BeautifulSoup object if successful, None otherwise
        """
        try:
            print(f"Fetching URL: {self.url}")
            response = requests.get(self.url, headers=self.headers, timeout=15)
            response.raise_for_status()
            print(f"Successfully fetched page (Status: {response.status_code})")
            return BeautifulSoup(response.content, 'lxml')
        except requests.RequestException as e:
            print(f"Error fetching page: {e}", file=sys.stderr)
            return None
    
    def extract_listings(self, soup: BeautifulSoup) -> List[Dict[str, str]]:
        """
        Extract listing data from the page.
        
        Args:
            soup: BeautifulSoup object of the page
            
        Returns:
            List of dictionaries containing listing data
        """
        listings = []
        
        try:
            # Strategy 1: Look for div elements with specific classes
            # Common patterns: listing-item, company-listing, yard-listing, etc.
            listing_containers = soup.find_all('div', class_=lambda x: x and (
                'listing' in str(x).lower() or 
                'yard' in str(x).lower() or 
                'company' in str(x).lower() or
                'item' in str(x).lower()
            ))
            
            # Strategy 2: If no divs found, try table rows
            if not listing_containers:
                listing_containers = soup.find_all('tr')[1:]  # Skip header row
            
            # Strategy 3: Try article or section tags
            if not listing_containers:
                listing_containers = soup.find_all(['article', 'section'])
            
            print(f"Found {len(listing_containers)} potential listing containers")
            
            for idx, container in enumerate(listing_containers, 1):
                try:
                    listing_data = self._extract_single_listing(container)
                    if listing_data:
                        listings.append(listing_data)
                        print(f"Extracted listing {idx}: {listing_data.get('Company', 'N/A')}")
                except Exception as e:
                    print(f"Error extracting listing {idx}: {e}", file=sys.stderr)
                    continue
            
            print(f"Successfully extracted {len(listings)} listings")
            
        except Exception as e:
            print(f"Error in extract_listings: {e}", file=sys.stderr)
        
        return listings
    
    def _extract_single_listing(self, container) -> Optional[Dict[str, str]]:
        """
        Extract data from a single listing container.
        
        Args:
            container: BeautifulSoup element containing listing data
            
        Returns:
            Dictionary with Company, Location, YardBrief, Address or None
        """
        listing = {
            'Company': '',
            'Location': '',
            'YardBrief': '',
            'Address': ''
        }
        
        # Extract company name
        # Try various selectors for company name
        company_elem = (
            container.find('h2') or 
            container.find('h3') or 
            container.find('a', class_=lambda x: x and 'company' in str(x).lower()) or
            container.find('a', class_=lambda x: x and 'name' in str(x).lower()) or
            container.find(class_=lambda x: x and 'company' in str(x).lower()) or
            container.find(class_=lambda x: x and 'name' in str(x).lower())
        )
        if company_elem:
            listing['Company'] = company_elem.get_text(strip=True)
        
        # Extract location
        location_elem = (
            container.find(class_=lambda x: x and 'location' in str(x).lower()) or
            container.find(class_=lambda x: x and 'city' in str(x).lower()) or
            container.find('span', class_=lambda x: x and 'location' in str(x).lower())
        )
        if location_elem:
            listing['Location'] = location_elem.get_text(strip=True)
        
        # Extract yard brief/description
        brief_elem = (
            container.find(class_=lambda x: x and 'brief' in str(x).lower()) or
            container.find(class_=lambda x: x and 'description' in str(x).lower()) or
            container.find(class_=lambda x: x and 'summary' in str(x).lower()) or
            container.find('p')
        )
        if brief_elem:
            listing['YardBrief'] = brief_elem.get_text(strip=True)
        
        # Extract full address
        address_elem = (
            container.find(class_=lambda x: x and 'address' in str(x).lower()) or
            container.find(class_=lambda x: x and 'addr' in str(x).lower())
        )
        if address_elem:
            listing['Address'] = address_elem.get_text(strip=True)
        else:
            # Try to construct address from multiple elements
            address_parts = []
            for elem in container.find_all(['span', 'div'], class_=lambda x: x and (
                'street' in str(x).lower() or 
                'city' in str(x).lower() or 
                'state' in str(x).lower() or
                'zip' in str(x).lower()
            )):
                text = elem.get_text(strip=True)
                if text and text not in address_parts:
                    address_parts.append(text)
            if address_parts:
                listing['Address'] = ', '.join(address_parts)
        
        # Only return listing if at least company name is found
        if listing['Company']:
            return listing
        
        return None
    
    def scrape(self) -> List[Dict[str, str]]:
        """
        Main scraping method.
        
        Returns:
            List of listing dictionaries
        """
        soup = self.fetch_page()
        if not soup:
            return []
        
        self.listings = self.extract_listings(soup)
        return self.listings
    
    def export_to_excel(self, filename: str = 'automotive_yards.xlsx'):
        """
        Export scraped data to Excel file.
        
        Args:
            filename: Output Excel filename
        """
        if not self.listings:
            print("No listings to export!", file=sys.stderr)
            return
        
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Automotive Yards"
            
            # Define headers
            headers = ['Company', 'Location', 'YardBrief', 'Address']
            
            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data
            for row_idx, listing in enumerate(self.listings, 2):
                ws.cell(row=row_idx, column=1, value=listing.get('Company', ''))
                ws.cell(row=row_idx, column=2, value=listing.get('Location', ''))
                ws.cell(row=row_idx, column=3, value=listing.get('YardBrief', ''))
                ws.cell(row=row_idx, column=4, value=listing.get('Address', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            print(f"Successfully exported {len(self.listings)} listings to {filename}")
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}", file=sys.stderr)


def main():
    """Main function."""
    url = "https://www.scrapmonster.com/scrap-yard/material/automotive/746"
    
    print("=" * 60)
    print("ScrapMonster Automotive Yard Scraper")
    print("=" * 60)
    
    scraper = ScrapMonsterScraper(url)
    listings = scraper.scrape()
    
    if listings:
        scraper.export_to_excel('automotive_yards.xlsx')
        print(f"\n{len(listings)} listings scraped successfully!")
    else:
        print("\nNo listings found or error occurred.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
